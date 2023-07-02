import openpyxl
from googletrans import Translator
import time
import TTS



def translate_excel (sheet):
    # Get the maximum row number in column A
    max_row = sheet.max_row

    # Create a translator object
    translator = Translator(service_urls=['translate.google.com'])

    count = 1
    for key in lang_code_dict:
        #decalring language code
        lang_code = lang_code_dict[key]
        print ("This is the language code" + str(lang_code))
        count = count + 1


        # Iterate through each cell in column A
        for row in range(1, max_row + 1):
            if row == 1:
                write_text = lang_code
            else:
                # Get the English word/phrase from column A
                english_text = sheet.cell(row=row, column=1).value
                print ("Row " + str(row) + " " + str(english_text)) 

                # Translate the text with retries
                max_retries = 5
                retry_delay = 1  # seconds

                for retry in range(max_retries):
                    try:
                        translation = translator.translate(english_text, dest=lang_code)
                        
                        break  # Translation successful, break out of the loop
                    except Exception as e:
                        print(f"Translation failed. Retrying... ({retry + 1}/{max_retries})")
                        time.sleep(retry_delay)

                # Handle translation result or error here
                if translation:
                    translated_text = translation.text
                    print (translated_text)
                else:
                    translated_text = "Translation failed"

                # Get the next empty row
                #snext_row = max_row + 1

                # Write the translation to the next empty row in column B
                write_text = translated_text
            sheet.cell(row=row, column=count).value = write_text
            print ("writing in row" + str(row)+ write_text)

            # Save the modified Excel file
            workbook.save('output_file.xlsx')

# Language Codes
lang_code_dict = {
        "Hindi": "hi",
        "Tamil": "ta",
        "Telugu": "te",
        "Kannada": "kn",
        "Marathi": "mr",
        "Malayalam": "ml",
        "Bengali": "bn",
        "Gujarati": "gu",
        "Punjabi": "pa",
        "Odia": "or"
    }

# Load the Excel file
workbook = openpyxl.load_workbook('input_file.xlsx')
sheet = workbook.active

translate_excel (sheet)
# Path to the Excel file
lang_file = 'output_file.xlsx'
TTS.TTS (lang_file)